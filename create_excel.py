#!/usr/bin/env python3
"""
Create Excel file from shares_outstanding.json
"""

import json
import pandas as pd
import sys

# Read JSON data
with open('shares_outstanding.json', 'r') as f:
    data = json.load(f)

# Convert to DataFrame
rows = []
for item in data['all_data']:
    rows.append({
        'Ticker': item['ticker'],
        'Date': item['date'],
        'Share Class Shares Outstanding': item.get('share_class_shares_outstanding'),
        'Weighted Shares Outstanding': item.get('weighted_shares_outstanding'),
        'Market Cap': item.get('market_cap'),
        'Company Name': item.get('name', '')
    })

df = pd.DataFrame(rows)

# Write to Excel
excel_file = 'shares_outstanding.xlsx'

try:
    # Try with openpyxl engine (most common)
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Shares Outstanding')
        worksheet = writer.sheets['Shares Outstanding']
        
        # Format header row
        from openpyxl.styles import Font, Alignment, PatternFill
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 12  # Ticker
        worksheet.column_dimensions['B'].width = 12  # Date
        worksheet.column_dimensions['C'].width = 30  # Share Class
        worksheet.column_dimensions['D'].width = 30  # Weighted
        worksheet.column_dimensions['E'].width = 20  # Market Cap
        worksheet.column_dimensions['F'].width = 50  # Company Name
        
        # Format numbers
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[2].value is not None:
                row[2].number_format = '#,##0'
            if row[3].value is not None:
                row[3].number_format = '#,##0'
            if row[4].value is not None:
                row[4].number_format = '#,##0.00'
            # Center align ticker and date
            row[0].alignment = Alignment(horizontal='center')
            row[1].alignment = Alignment(horizontal='center')
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"✓ Excel file created: {excel_file}")
    print(f"  Rows: {len(df)}")
    
except ImportError as e:
    print(f"Error: Missing required library. Please install openpyxl:", file=sys.stderr)
    print(f"  pip install openpyxl", file=sys.stderr)
    sys.exit(1)
except Exception as e:
    print(f"Error creating Excel file: {e}", file=sys.stderr)
    sys.exit(1)
