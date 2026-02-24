#!/usr/bin/env python3
"""
Reorganize shares outstanding data into pivot format:
- Rows: Tickers
- Columns: Dates
"""

import json
import pandas as pd
import sys
from openpyxl.styles import Font, Alignment, PatternFill

# Read JSON data
with open('shares_outstanding.json', 'r') as f:
    data = json.load(f)

# Convert to DataFrame in long format first
# Filter for Class A shares only
rows = []
for item in data['all_data']:
    name = item.get('name', '') or ''
    # Only include Class A shares (check uppercase)
    if 'CLASS A' in name.upper():
        rows.append({
            'Ticker': item['ticker'],
            'Date': item['date'],
            'Share Class Shares Outstanding': item.get('share_class_shares_outstanding'),
            'Weighted Shares Outstanding': item.get('weighted_shares_outstanding'),
            'Market Cap': item.get('market_cap'),
        })

if not rows:
    print("No Class A shares found in data. Please re-run fetch_shares_outstanding.py", file=sys.stderr)
    sys.exit(1)

df_long = pd.DataFrame(rows)

# Pivot to wide format - using Share Class Shares Outstanding
df_pivot = df_long.pivot_table(
    index='Ticker',
    columns='Date',
    values='Share Class Shares Outstanding',
    aggfunc='first'
)

# Reset index to make Ticker a column
df_pivot = df_pivot.reset_index()

# Rename columns - format dates as requested (M/D/YYYY)
date_columns = []
for col in df_pivot.columns:
    if col == 'Ticker':
        date_columns.append('Ticker')
    else:
        # Parse date and reformat
        from datetime import datetime
        dt = datetime.strptime(col, '%Y-%m-%d')
        formatted_date = dt.strftime('%-m/%-d/%Y')  # Remove leading zeros
        date_columns.append(formatted_date)

df_pivot.columns = date_columns

# Reorder columns to match user's example (Ticker first, then dates)
ticker_col = df_pivot.pop('Ticker')
df_pivot.insert(0, 'Ticker', ticker_col)

# Sort by ticker
df_pivot = df_pivot.sort_values('Ticker').reset_index(drop=True)

# Write CSV
csv_file = 'shares_outstanding.csv'
with open(csv_file, 'w') as f:
    # Write header
    f.write('As of 13-F Filing Date\t\t\t\t\n')
    f.write('\t'.join(df_pivot.columns) + '\n')
    
    # Write data rows
    for _, row in df_pivot.iterrows():
        values = []
        for col in df_pivot.columns:
            val = row[col]
            if col == 'Ticker':
                values.append(str(val))
            elif pd.isna(val):
                values.append('N/A')
            else:
                # Format as number with commas
                values.append(f"{int(val):,}")
        f.write('\t'.join(values) + '\n')

print(f"✓ CSV file updated: {csv_file}")

# Write Excel
excel_file = 'shares_outstanding.xlsx'
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df_pivot.to_excel(writer, index=False, sheet_name='Shares Outstanding', startrow=1)
    worksheet = writer.sheets['Shares Outstanding']
    
    # Add header row "As of 13-F Filing Date"
    worksheet.merge_cells('A1:E1')
    header_cell = worksheet['A1']
    header_cell.value = 'As of 13-F Filing Date'
    header_cell.font = Font(bold=True, size=12)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Format column headers (row 2)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 12  # Ticker
    for col_idx, col in enumerate(df_pivot.columns[1:], start=2):  # Date columns
        col_letter = worksheet.cell(row=2, column=col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = 18
    
    # Format numbers in data rows
    for row_idx in range(3, worksheet.max_row + 1):
        for col_idx in range(2, len(df_pivot.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value) != 'nan':
                try:
                    cell.number_format = '#,##0'
                except:
                    pass
        # Center align ticker column
        worksheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
    
    # Freeze header rows
    worksheet.freeze_panes = 'A3'

print(f"✓ Excel file updated: {excel_file}")
print(f"  Rows: {len(df_pivot)}")
print(f"  Columns: {len(df_pivot.columns)}")
